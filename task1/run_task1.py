"""
任务一：构建结构化财报数据库
=================================
功能：
1. 扫描所有PDF财报文件
2. 提取文本和表格数据
3. 规则提取 + LLM增强提取财务数据
4. 数据校验和清洗
5. 存入SQLite结构化数据库
6. 导入公司基本信息
7. 打印详细处理进度

使用方法：
    python task1/run_task1.py
"""
import sys
import os
import asyncio
import json
import logging
import time
from collections import defaultdict
from pathlib import Path
from datetime import datetime

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(PROJECT_ROOT))

from backend.core.config import AppConfig, DATA_DIR, RESULTS_DIR
from backend.core.database import DatabaseManager
from backend.core.llm_client import LLMClient
from backend.core.preflight import emit_preflight_report, run_task1_preflight
from backend.core.pdf_parser import (
    scan_report_files, extract_text_from_pdf, extract_tables_from_pdf,
    classify_report_by_content, extract_financial_data_by_rules,
    ReportMeta, FinancialData,
)

# =============================================================================
# 日志配置
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(PROJECT_ROOT / "logs" / "task1.log"), encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# 公司信息导入
# =============================================================================
def import_company_info(db: DatabaseManager, data_dir: Path):
    """导入附件1的公司基本信息"""
    import openpyxl
    
    xlsx_files = []
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and ('1' in f or '基本信息' in f or '公司' in f):
            xlsx_files.append(data_dir / f)
    
    if not xlsx_files:
        # 尝试找到附件1
        for f in os.listdir(data_dir):
            fp = data_dir / f
            if f.endswith('.xlsx') and fp.is_file():
                try:
                    wb = openpyxl.load_workbook(str(fp))
                    if len(wb.sheetnames) == 2:  # 附件1有2个sheet
                        xlsx_files.append(fp)
                        break
                except Exception:
                    pass
    
    if not xlsx_files:
        logger.warning("未找到公司基本信息文件(附件1)")
        return
    
    xlsx_path = str(xlsx_files[0])
    logger.info(f"导入公司基本信息: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]  # 第一个sheet是基本信息表
    
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        logger.warning("公司信息表数据为空")
        return
    
    headers = [str(h) for h in rows[0]]
    records = []
    
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        
        # 映射到数据库字段
        db_record = {
            "serial_number": record.get("序号"),
            "stock_code": str(record.get("股票代码", "")),
            "stock_abbr": record.get("A股简称", ""),
            "company_name": record.get("公司名称", ""),
            "english_name": record.get("英文名称", ""),
            "industry": record.get("所属证监会行业", ""),
            "exchange": record.get("上市交易所", ""),
            "security_type": record.get("证券类别", ""),
            "registered_area": record.get("注册区域", ""),
            "registered_capital": str(record.get("注册资本", "")),
            "employee_count": record.get("雇员人数"),
            "management_count": record.get("管理人员人数"),
        }
        records.append(db_record)
        logger.info(f"  公司: {db_record['stock_abbr']} ({db_record['stock_code']})")
    
    db.insert_many("company_info", records)
    logger.info(f"成功导入 {len(records)} 家公司信息")
    return records


# =============================================================================
# 构建股票代码→简称映射
# =============================================================================
def build_stock_mapping(db: DatabaseManager) -> dict:
    """从数据库获取股票代码到简称的映射"""
    rows = db.execute_query("SELECT stock_code, stock_abbr FROM company_info")
    mapping = {}
    for row in rows:
        mapping[str(row['stock_code'])] = row['stock_abbr']
    return mapping


def clone_report_meta(meta: ReportMeta) -> ReportMeta:
    """复制报告元数据，避免缓存对象被后续流程意外修改。"""
    return ReportMeta(**vars(meta))


def build_report_group_indices(reports: list):
    """构建按发布日期和报告期的报告索引。"""
    by_publish_date = defaultdict(list)
    by_report_period = defaultdict(list)

    for report in reports:
        identities = {str(v) for v in [report.stock_code, report.stock_abbr] if v}
        for identity in identities:
            if report.publish_date:
                by_publish_date[(identity, report.publish_date)].append(report)
            if report.report_period:
                by_report_period[(identity, report.report_period)].append(report)

    return by_publish_date, by_report_period


# =============================================================================
# LLM增强提取
# =============================================================================
LLM_EXTRACT_PROMPT = """
你是一个专业的财务数据提取专家。请从以下财务报告文本中提取指定的财务数据。

## 要提取的数据字段

### 利润表字段（单位：万元，除特别标注外）
- net_profit: 净利润(万元)
- total_operating_revenue: 营业总收入(万元)
- operating_expense_cost_of_sales: 营业成本(万元)
- operating_expense_selling_expenses: 销售费用(万元)
- operating_expense_administrative_expenses: 管理费用(万元)
- operating_expense_financial_expenses: 财务费用(万元)
- operating_expense_rnd_expenses: 研发费用(万元)
- operating_expense_taxes_and_surcharges: 税金及附加(万元)
- total_operating_expenses: 营业总支出(万元)
- operating_profit: 营业利润(万元)
- total_profit: 利润总额(万元)
- other_income: 其他收益(万元)
- investment_income: 投资收益(万元)
- fair_value_change_income: 公允价值变动收益(万元)
- asset_disposal_income: 资产处置收益(万元)
- non_operating_income: 营业外收入(万元)
- non_operating_expenses: 营业外支出(万元)
- asset_impairment_loss: 资产减值损失(万元，损失为负数)
- credit_impairment_loss: 信用减值损失(万元，损失为负数)

### 资产负债表字段（单位：万元）
- asset_cash_and_cash_equivalents: 货币资金(万元)
- asset_accounts_receivable: 应收账款(万元)
- asset_inventory: 存货(万元)
- asset_trading_financial_assets: 交易性金融资产(万元)
- asset_construction_in_progress: 在建工程(万元)
- asset_total_assets: 总资产(万元)
- liability_accounts_payable: 应付账款(万元)
- liability_advance_from_customers: 预收账款(万元)
- liability_total_liabilities: 总负债(万元)
- liability_contract_liabilities: 合同负债(万元)
- liability_short_term_loans: 短期借款(万元)
- equity_unappropriated_profit: 未分配利润(万元)
- equity_total_equity: 股东权益合计(万元)

### 现金流量表字段（单位：万元，除特别标注外）
- net_cash_flow: 现金及现金等价物净增加额(元，注意单位是元不是万元)
- operating_cf_net_amount: 经营活动现金流量净额(万元)
- operating_cf_cash_from_sales: 销售商品收到的现金(万元)
- investing_cf_net_amount: 投资活动现金流量净额(万元)
- investing_cf_cash_for_investments: 投资支付的现金(万元)
- investing_cf_cash_from_investment_recovery: 收回投资收到的现金(万元)
- financing_cf_cash_from_borrowing: 取得借款收到的现金(万元)
- financing_cf_cash_for_debt_repayment: 偿还债务支付的现金(万元)
- financing_cf_net_amount: 融资活动现金流量净额(万元)

### 核心业绩指标
- eps: 基本每股收益(元)
- net_asset_per_share: 每股净资产(元)
- roe: 净资产收益率(%)
- operating_cf_per_share: 每股经营现金流量(元)
- gross_profit_margin: 销售毛利率(%)

## 重要提示
1. 所有金额字段单位为万元（除了 eps, net_asset_per_share, operating_cf_per_share 单位为元，以及 net_cash_flow 单位为元）
2. 如果原文中金额单位是"元"，需要除以10000转换为万元
3. 如果找不到某个字段，设置为null
4. 百分比字段直接填数字（如15.5表示15.5%）
5. 负数用负号表示

## 报告信息
- 公司：{stock_abbr} ({stock_code})
- 报告期：{report_period}
- 报告类型：{report_type}

## 财报文本内容（部分）
{text_content}

请返回JSON格式：
{{
    "income_sheet": {{...}},
    "balance_sheet": {{...}},
    "cash_flow_sheet": {{...}},
    "core_performance": {{...}}
}}
"""


async def llm_enhance_extraction(
    llm: LLMClient,
    text: str,
    meta: ReportMeta,
    rule_data: FinancialData,
) -> FinancialData:
    """使用LLM增强数据提取"""
    # 截取关键部分（避免超长）
    text_truncated = text[:12000]
    
    prompt = LLM_EXTRACT_PROMPT.format(
        stock_abbr=meta.stock_abbr,
        stock_code=meta.stock_code,
        report_period=meta.report_period,
        report_type=meta.report_type,
        text_content=text_truncated,
    )
    
    try:
        result = await llm.query_json(prompt, temperature=0.1)
        
        # 合并结果：LLM提取的数据覆盖规则提取的空值
        for sheet_key in ["income_sheet", "balance_sheet", "cash_flow_sheet", "core_performance"]:
            llm_data = result.get(sheet_key, {})
            rule_sheet = getattr(rule_data, sheet_key, {})
            
            for k, v in llm_data.items():
                if v is not None and k not in rule_sheet:
                    rule_sheet[k] = v
            
            setattr(rule_data, sheet_key, rule_sheet)
        
        logger.info(f"  LLM增强提取完成: {meta.file_name}")
    except Exception as e:
        logger.warning(f"  LLM增强提取失败: {e}")
    
    return rule_data


# =============================================================================
# 数据校验
# =============================================================================
def validate_financial_data(data: FinancialData) -> list:
    """
    多维度数据校验
    返回校验问题列表
    """
    issues = []
    meta = data.meta
    prefix = f"[{meta.stock_abbr} {meta.report_period}]"
    
    income = data.income_sheet
    balance = data.balance_sheet
    cash_flow = data.cash_flow_sheet
    
    # 1. 利润表校验：使用中国会计准则正确公式
    # 营业利润 = 营业总收入 - 营业总成本 + 其他收益 + 投资收益 + 公允价值变动收益
    #            + 信用减值损失(已含符号) + 资产减值损失(已含符号) + 资产处置收益
    rev = income.get("total_operating_revenue")
    exp = income.get("total_operating_expenses")
    op_profit = income.get("operating_profit")
    if rev is not None and exp is not None and op_profit is not None:
        calc_profit = rev - exp
        # 加上已提取的中间项（中国会计准则中这些项目介于营业总成本和营业利润之间）
        other_items = {
            "other_income": income.get("other_income"),
            "investment_income": income.get("investment_income"),
            "fair_value_change_income": income.get("fair_value_change_income"),
            "credit_impairment_loss": income.get("credit_impairment_loss"),
            "asset_impairment_loss": income.get("asset_impairment_loss"),
            "asset_disposal_income": income.get("asset_disposal_income"),
        }
        extracted_count = 0
        for k, v in other_items.items():
            if v is not None:
                calc_profit += v
                extracted_count += 1
        
        diff = abs(calc_profit - op_profit)
        total_items = len(other_items)  # 6个中间项
        missing_count = total_items - extracted_count
        
        # 根据中间项覆盖率动态调整容差
        # 缺少的中间项越多，容差越大；全部缺失时跳过校验
        if extracted_count == 0:
            threshold = None  # 无中间项时跳过校验（公式不完整，无法可靠校验）
        elif missing_count == 0:
            threshold = abs(op_profit) * 0.05 + 100  # 全部提取到，严格校验
        elif missing_count <= 2:
            threshold = abs(op_profit) * 0.5 + abs(rev) * 0.15 + 5000  # 缺少1-2项，放宽（缺失项可能数千万）
        else:
            threshold = abs(op_profit) * 1.0 + abs(rev) * 0.2 + 10000  # 缺少3+项，大幅放宽
        
        if threshold is not None and diff > threshold:
            extracted_items = {k: v for k, v in other_items.items() if v is not None}
            issues.append(
                f"{prefix} 利润表校验({extracted_count}/{total_items}项): "
                f"营业总收入({rev})-营业总成本({exp})+中间项={calc_profit}, "
                f"但营业利润={op_profit}, 差额={diff}"
            )
    
    # 1b. 利润总额校验：利润总额 = 营业利润 + 营业外收入 - 营业外支出
    total_profit = income.get("total_profit")
    non_op_income = income.get("non_operating_income")
    non_op_expense = income.get("non_operating_expenses")
    if op_profit is not None and total_profit is not None:
        if non_op_income is not None and non_op_expense is not None:
            calc_total = op_profit + non_op_income - non_op_expense
            diff = abs(calc_total - total_profit)
            if diff > abs(total_profit) * 0.05 + 50:
                issues.append(
                    f"{prefix} 利润总额校验: 营业利润({op_profit})+营业外收入({non_op_income})"
                    f"-营业外支出({non_op_expense})={calc_total}, 但利润总额={total_profit}, 差额={diff}"
                )
    
    # 2. 资产负债表校验：总资产 = 总负债 + 股东权益
    total_assets = balance.get("asset_total_assets")
    total_liab = balance.get("liability_total_liabilities")
    total_equity = balance.get("equity_total_equity")
    if total_assets is not None and total_liab is not None and total_equity is not None:
        calc_assets = total_liab + total_equity
        diff = abs(calc_assets - total_assets)
        if diff > abs(total_assets) * 0.01 + 10:  # 允许1%误差或10万元
            issues.append(
                f"{prefix} 资产负债表校验: 总负债({total_liab})+股东权益({total_equity})={calc_assets}, "
                f"但总资产={total_assets}, 差额={diff}"
            )
    
    # 3. 资产负债率校验
    if total_assets is not None and total_liab is not None and total_assets != 0:
        calc_ratio = total_liab / total_assets * 100
        reported_ratio = balance.get("asset_liability_ratio")
        if reported_ratio is not None:
            diff = abs(calc_ratio - reported_ratio)
            if diff > 1:  # 允许1个百分点误差
                issues.append(
                    f"{prefix} 资产负债率校验: 计算值={calc_ratio:.2f}%, 报告值={reported_ratio}%"
                )
    
    # 4. 数据完整性检查
    required_income_fields = ["total_operating_revenue", "net_profit", "total_profit"]
    for field in required_income_fields:
        if income.get(field) is None:
            issues.append(f"{prefix} 利润表缺失关键字段: {field}")
    
    required_balance_fields = ["asset_total_assets", "liability_total_liabilities", "equity_total_equity"]
    for field in required_balance_fields:
        if balance.get(field) is None:
            issues.append(f"{prefix} 资产负债表缺失关键字段: {field}")
    
    # 5. 数值合理性检查
    if income.get("net_profit") is not None and income.get("total_operating_revenue") is not None:
        if income["total_operating_revenue"] != 0:
            margin = income["net_profit"] / income["total_operating_revenue"] * 100
            if abs(margin) > 100:
                issues.append(
                    f"{prefix} 净利率异常: {margin:.1f}% (净利润/营业收入)"
                )
    
    return issues


async def llm_validate_data(llm: LLMClient, data: FinancialData, issues: list) -> list:
    """使用LLM辅助校验"""
    if not issues:
        return issues
    
    prompt = f"""
以下是从财务报告中提取的数据校验结果，请分析这些问题并给出建议：

公司: {data.meta.stock_abbr} ({data.meta.stock_code})
报告期: {data.meta.report_period}

校验发现的问题:
{json.dumps(issues, ensure_ascii=False, indent=2)}

提取的利润表数据:
{json.dumps(data.income_sheet, ensure_ascii=False, indent=2)}

提取的资产负债表数据:
{json.dumps(data.balance_sheet, ensure_ascii=False, indent=2)}

请分析：
1. 哪些问题是真正的数据错误？
2. 哪些可能是因为报告格式导致的提取误差？
3. 建议如何修正？

返回JSON格式：
{{
    "real_issues": ["真正需要关注的问题"],
    "corrections": {{"字段名": "建议修正值"}},
    "notes": "其他说明"
}}
"""
    try:
        result = await llm.query_json(prompt, temperature=0.1)
        return result.get("real_issues", issues)
    except Exception:
        return issues


# =============================================================================
# 数据入库
# =============================================================================
def save_to_database(db: DatabaseManager, data: FinancialData):
    """将提取的财务数据存入数据库"""
    meta = data.meta
    replace_existing = "摘要" not in meta.report_type
    
    common_fields = {
        "stock_code": meta.stock_code,
        "stock_abbr": meta.stock_abbr,
        "report_period": meta.report_period,
        "report_year": meta.report_year,
    }
    
    # 利润表
    if data.income_sheet:
        record = {**common_fields, **data.income_sheet}
        record.pop("serial_number", None)  # 自增
        db.insert_record("income_sheet", record, replace=replace_existing)
    
    # 资产负债表
    if data.balance_sheet:
        record = {**common_fields, **data.balance_sheet}
        record.pop("serial_number", None)
        db.insert_record("balance_sheet", record, replace=replace_existing)
    
    # 现金流量表
    if data.cash_flow_sheet:
        record = {**common_fields, **data.cash_flow_sheet}
        record.pop("serial_number", None)
        db.insert_record("cash_flow_sheet", record, replace=replace_existing)
    
    # 核心业绩指标
    if data.core_performance:
        # 补充利润表中的指标
        perf = {**common_fields, **data.core_performance}
        if "total_operating_revenue" not in perf and "total_operating_revenue" in data.income_sheet:
            perf["total_operating_revenue"] = data.income_sheet["total_operating_revenue"]
        if "net_profit_10k_yuan" not in perf and "net_profit" in data.income_sheet:
            perf["net_profit_10k_yuan"] = data.income_sheet["net_profit"]
        perf.pop("serial_number", None)
        db.insert_record("core_performance_indicators_sheet", perf, replace=replace_existing)
    
    logger.info(f"  数据已入库: {meta.stock_abbr} {meta.report_period}")


# =============================================================================
# 计算同比/环比增长率
# =============================================================================
def calculate_growth_rates(db: DatabaseManager):
    """在所有数据入库后，计算同比和环比增长率"""
    logger.info("\n" + "=" * 60)
    logger.info("计算同比和环比增长率...")
    
    # 获取所有公司和报告期
    periods = db.execute_query("""
        SELECT DISTINCT stock_code, stock_abbr, report_period, report_year 
        FROM income_sheet ORDER BY stock_code, report_period
    """)
    
    for p in periods:
        code = p['stock_code']
        period = p['report_period']
        year = p['report_year']
        
        # 确定同比对应期（去年同期）
        suffix = period.replace(str(year), "")
        yoy_period = f"{year - 1}{suffix}"
        
        # 利润表同比
        current = db.execute_query(
            "SELECT * FROM income_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        prev = db.execute_query(
            "SELECT * FROM income_sheet WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        
        if current and prev:
            curr_rev = current[0].get("total_operating_revenue")
            prev_rev = prev[0].get("total_operating_revenue")
            curr_profit = current[0].get("net_profit")
            prev_profit = prev[0].get("net_profit")
            
            updates = {}
            if curr_rev is not None and prev_rev is not None and prev_rev != 0:
                updates["operating_revenue_yoy_growth"] = round((curr_rev - prev_rev) / abs(prev_rev) * 100, 4)
            if curr_profit is not None and prev_profit is not None and prev_profit != 0:
                updates["net_profit_yoy_growth"] = round((curr_profit - prev_profit) / abs(prev_profit) * 100, 4)
            
            if updates:
                set_clause = ", ".join([f"{k}=?" for k in updates.keys()])
                db.execute_sql(
                    f"UPDATE income_sheet SET {set_clause} WHERE stock_code=? AND report_period=?",
                    tuple(updates.values()) + (code, period)
                )
                
                # 同步更新核心业绩表
                for k, v in updates.items():
                    if k in ["operating_revenue_yoy_growth", "net_profit_yoy_growth"]:
                        db.execute_sql(
                            f"UPDATE core_performance_indicators_sheet SET {k}=? WHERE stock_code=? AND report_period=?",
                            (v, code, period)
                        )
        
        # 资产负债表同比
        current_bs = db.execute_query(
            "SELECT * FROM balance_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        prev_bs = db.execute_query(
            "SELECT * FROM balance_sheet WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        
        if current_bs and prev_bs:
            updates = {}
            for field, growth_field in [
                ("asset_total_assets", "asset_total_assets_yoy_growth"),
                ("liability_total_liabilities", "liability_total_liabilities_yoy_growth"),
            ]:
                curr_val = current_bs[0].get(field)
                prev_val = prev_bs[0].get(field)
                if curr_val is not None and prev_val is not None and prev_val != 0:
                    updates[growth_field] = round((curr_val - prev_val) / abs(prev_val) * 100, 4)
            
            if updates:
                set_clause = ", ".join([f"{k}=?" for k in updates.keys()])
                db.execute_sql(
                    f"UPDATE balance_sheet SET {set_clause} WHERE stock_code=? AND report_period=?",
                    tuple(updates.values()) + (code, period)
                )
    
    logger.info("同比/环比增长率计算完成")


# =============================================================================
# 主流程
# =============================================================================
async def main():
    start_time = time.time()
    
    print("=" * 70)
    print("  任务一：构建结构化财报数据库")
    print("  开始时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 70)
    
    # 1. 加载配置
    config = AppConfig.load()
    active_data_dir = config.get_sample_data_dir()
    print(f"\n[1/8] 加载配置完成")
    print(f"  LLM API: {[c.name for c in config.get_enabled_llms()]}")
    print(f"  数据目录: {active_data_dir}")
    preflight = run_task1_preflight(config, active_data_dir)
    emit_preflight_report(preflight, logger_instance=logger)
    if not preflight.ok:
        raise SystemExit("任务一运行前自检未通过，请先处理上述错误。")
    
    # 2. 初始化数据库
    db = DatabaseManager(config.db_path)
    db.init_db()
    print(f"\n[2/8] 数据库初始化完成: {config.db_path}")
    
    # 3. 导入公司基本信息
    print(f"\n[3/8] 导入公司基本信息...")
    stock_mapping = {}
    try:
        records = import_company_info(db, active_data_dir)
        if records:
            for r in records:
                stock_mapping[str(r['stock_code'])] = r['stock_abbr']
                print(f"  ✓ {r['stock_abbr']} (代码: {r['stock_code']})")
    except Exception as e:
        logger.error(f"导入公司信息失败: {e}")
    
    # 如果从数据库获取映射
    if not stock_mapping:
        stock_mapping = build_stock_mapping(db)
    
    # 4. 扫描财报文件
    print(f"\n[4/8] 扫描财报文件...")
    reports = scan_report_files(str(active_data_dir))
    print(f"  共发现 {len(reports)} 个PDF文件")
    reports_by_publish_date, reports_by_report_period = build_report_group_indices(reports)
    classified_meta_cache = {}
    
    # 按公司分组显示
    by_company = {}
    for r in reports:
        key = r.stock_abbr or r.stock_code or "未知"
        by_company.setdefault(key, []).append(r)
    
    for company, reps in by_company.items():
        print(f"  - {company}: {len(reps)} 份报告")
    
    # 5. 初始化LLM客户端
    print(f"\n[5/8] 初始化LLM客户端...")
    llm = LLMClient(config)
    print(f"  已准备 {len(llm.clients)} 个LLM API")

    def get_classified_meta(report_meta: ReportMeta, text: str = None) -> ReportMeta:
        cache_key = report_meta.file_path or report_meta.file_name
        cached = classified_meta_cache.get(cache_key)
        if cached is not None:
            return cached

        candidate = clone_report_meta(report_meta)
        if text is not None:
            candidate = classify_report_by_content(text, candidate)
        elif candidate.report_type == "待确定" or not candidate.report_period:
            candidate_text = extract_text_from_pdf(candidate.file_path)
            candidate = classify_report_by_content(candidate_text, candidate)

        if not candidate.stock_abbr and candidate.stock_code in stock_mapping:
            candidate.stock_abbr = stock_mapping[candidate.stock_code]

        identities = {str(v) for v in [candidate.stock_code, candidate.stock_abbr] if v}
        for identity in identities:
            if candidate.report_period:
                group_key = (identity, candidate.report_period)
                if report_meta not in reports_by_report_period[group_key]:
                    reports_by_report_period[group_key].append(report_meta)

        classified_meta_cache[cache_key] = candidate
        return candidate

    def has_full_report_counterpart(current_meta: ReportMeta) -> bool:
        if not current_meta.report_period:
            return False

        candidate_reports = []
        seen_paths = set()
        identities = {str(v) for v in [current_meta.stock_code, current_meta.stock_abbr] if v}

        for identity in identities:
            if current_meta.publish_date:
                for report in reports_by_publish_date.get((identity, current_meta.publish_date), []):
                    if report.file_path not in seen_paths:
                        candidate_reports.append(report)
                        seen_paths.add(report.file_path)
            for report in reports_by_report_period.get((identity, current_meta.report_period), []):
                if report.file_path not in seen_paths:
                    candidate_reports.append(report)
                    seen_paths.add(report.file_path)

        for report in candidate_reports:
            if report.file_path == current_meta.file_path:
                continue
            candidate_meta = get_classified_meta(report)
            if candidate_meta.report_period == current_meta.report_period and "摘要" not in candidate_meta.report_type:
                return True

        return False
    
    # 6. 逐个处理财报
    print(f"\n[6/8] 开始解析财报...")
    total = len(reports)
    success_count = 0
    fail_count = 0
    all_issues = []
    
    # 限制并发数
    semaphore = asyncio.Semaphore(3)
    
    async def process_single_report(idx, meta):
        nonlocal success_count, fail_count
        
        async with semaphore:
            prefix = f"  [{idx+1}/{total}]"
            
            try:
                # 只有确认存在同报告期完整版时，才提前跳过摘要。
                if "摘要" in meta.report_type:
                    meta = get_classified_meta(meta)
                    if has_full_report_counterpart(meta):
                        logger.info(f"{prefix} 检测到同报告期完整版，跳过报告摘要: {meta.file_name}")
                        print(f"{prefix}   跳过: {meta.report_type}(已存在完整版)")
                        return

                print(f"{prefix} 处理: {meta.file_name}")
                
                # Step A: 提取PDF文本
                text = extract_text_from_pdf(meta.file_path)
                print(f"{prefix}   文本提取完成, 长度: {len(text)} 字符")
                
                # Step B: 补充元数据
                meta = get_classified_meta(meta, text=text)
                
                # 补充股票简称
                if not meta.stock_abbr and meta.stock_code in stock_mapping:
                    meta.stock_abbr = stock_mapping[meta.stock_code]

                # 内容识别为摘要时，只在确认存在完整版的情况下跳过。
                # 如果当前期数只有摘要，则保留摘要作为兜底数据源。
                if "摘要" in meta.report_type and has_full_report_counterpart(meta):
                    logger.info(f"{prefix} 内容识别为报告摘要，且已检测到完整版，跳过入库: {meta.file_name}")
                    print(f"{prefix}   跳过: {meta.report_type}(已存在完整版)")
                    return
                
                if not meta.report_period:
                    logger.warning(f"{prefix}   无法确定报告期，跳过")
                    fail_count += 1
                    return
                
                print(f"{prefix}   报告期: {meta.report_period}, 类型: {meta.report_type}")
                
                # Step C: 提取表格
                tables = extract_tables_from_pdf(meta.file_path)
                print(f"{prefix}   表格提取完成, 共 {len(tables)} 个表格")
                
                # Step D: 规则提取
                data = extract_financial_data_by_rules(text, tables, meta)
                rule_fields = sum(len(v) for v in [data.income_sheet, data.balance_sheet, data.cash_flow_sheet, data.core_performance] if v)
                print(f"{prefix}   规则提取完成, 提取 {rule_fields} 个字段")
                
                # Step E: LLM增强提取（仅对年报和半年报使用，数据量大）
                if meta.report_type in ["年度报告", "半年度报告"] or rule_fields < 10:
                    print(f"{prefix}   调用LLM增强提取...")
                    data = await llm_enhance_extraction(llm, text, meta, data)
                    enhanced_fields = sum(len(v) for v in [data.income_sheet, data.balance_sheet, data.cash_flow_sheet, data.core_performance] if v)
                    print(f"{prefix}   LLM增强后共 {enhanced_fields} 个字段")
                
                # Step F: 数据校验
                issues = validate_financial_data(data)
                if issues:
                    print(f"{prefix}   ⚠ 发现 {len(issues)} 个校验问题:")
                    for issue in issues:
                        print(f"{prefix}     - {issue}")
                    all_issues.extend(issues)
                else:
                    print(f"{prefix}   ✓ 数据校验通过")
                
                # Step G: 入库
                save_to_database(db, data)
                success_count += 1
                print(f"{prefix}   ✓ 入库完成")
                
            except Exception as e:
                fail_count += 1
                logger.error(f"{prefix} 处理失败: {e}", exc_info=True)
                print(f"{prefix}   ✗ 处理失败: {e}")
    
    # 并发处理所有报告
    tasks = [process_single_report(i, meta) for i, meta in enumerate(reports)]
    await asyncio.gather(*tasks)
    
    # 7. 计算增长率
    print(f"\n[7/8] 计算同比/环比增长率...")
    calculate_growth_rates(db)
    
    # 8. 汇总报告
    elapsed = time.time() - start_time
    print(f"\n[8/8] 处理完成!")
    print("=" * 70)
    print(f"  总计文件: {total}")
    print(f"  成功处理: {success_count}")
    print(f"  跳过/失败: {fail_count}")
    print(f"  校验问题: {len(all_issues)}")
    print(f"  耗时: {elapsed:.1f} 秒")
    print()
    
    # 打印数据库统计
    print("数据库统计:")
    for table in ["company_info", "core_performance_indicators_sheet", "balance_sheet", "income_sheet", "cash_flow_sheet"]:
        count = db.get_table_row_count(table)
        print(f"  {table}: {count} 条记录")
    
    # 保存校验报告
    if all_issues:
        report_path = str(PROJECT_ROOT / "logs" / "validation_report.txt")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("数据校验报告\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            for issue in all_issues:
                f.write(f"- {issue}\n")
        print(f"\n校验报告已保存: {report_path}")
    
    print("\n" + "=" * 70)
    print("  任务一完成！结构化财报数据库已构建。")
    print("=" * 70)


if __name__ == "__main__":
    # 确保日志目录存在
    os.makedirs(str(PROJECT_ROOT / "logs"), exist_ok=True)
    os.makedirs(str(DATA_DIR), exist_ok=True)
    os.makedirs(str(RESULTS_DIR), exist_ok=True)
    
    asyncio.run(main())

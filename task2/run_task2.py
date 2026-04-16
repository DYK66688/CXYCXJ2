"""
任务二：搭建"智能问数"助手并回答附件4的问题
=================================
功能：
1. 加载任务一构建的结构化数据库
2. 初始化智能问答Agent
3. 逐个处理附件4中的问题（多轮对话）
4. 生成可视化图表到results目录
5. 输出result_2.xlsx提交文件

使用方法：
    python task2/run_task2.py
"""
import sys
import os
import asyncio
import json
import logging
import time
from pathlib import Path
from datetime import datetime

PROJECT_ROOT = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(PROJECT_ROOT))

from backend.core.config import AppConfig, SAMPLE_DATA_DIR, DATA_DIR, RESULTS_DIR
from backend.core.database import DatabaseManager
from backend.core.llm_client import LLMClient
from backend.core.preflight import emit_preflight_report, run_task2_preflight
from backend.core.agent import SmartQAAgent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(PROJECT_ROOT / "logs" / "task2.log"), encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


def load_questions(data_dir: Path) -> list:
    """加载附件4的问题"""
    import openpyxl
    
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and ('4' in f or '问题' in f):
            fp = data_dir / f
            try:
                wb = openpyxl.load_workbook(str(fp))
                if len(wb.sheetnames) == 1:
                    ws = wb[wb.sheetnames[0]]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) >= 2 and rows[0][0] and '编号' in str(rows[0][0]):
                        # 检查是否是附件4（B1xxx系列）
                        if 'B1' in str(rows[1][0]):
                            questions = []
                            for row in rows[1:]:
                                questions.append({
                                    "id": str(row[0]),
                                    "type": str(row[1]),
                                    "questions": json.loads(str(row[2])),
                                })
                            return questions
            except Exception as e:
                logger.warning(f"读取 {f} 失败: {e}")
    
    logger.warning("未找到附件4的问题文件")
    return []


def _extract_results_from_answers(answers: list) -> dict:
    """
    从agent返回的answers列表中提取SQL、图表类型、图片路径
    answers格式: [{"Q": "...", "A": {"content": "...", "image": [...]}, "sql": "...", "chart_type": "..."}, ...]
    """
    all_sql = []
    all_chart_types = []
    all_images = []

    for a in answers:
        if not isinstance(a, dict):
            continue
        # SQL在answer的顶层
        if a.get("sql"):
            all_sql.append(a["sql"])
        # chart_type在answer的顶层
        if a.get("chart_type"):
            all_chart_types.append(a["chart_type"])
        # 图片在 A.image 中
        answer_body = a.get("A", {})
        if isinstance(answer_body, dict) and answer_body.get("image"):
            for img_path in answer_body["image"]:
                # 只保留文件名
                img_name = os.path.basename(str(img_path))
                all_images.append(img_name)

    # 合并SQL（多轮对话可能有多条）
    sql_text = ";\n".join(all_sql) if all_sql else ""
    # 图表类型取所有出现的类型
    chart_type_text = ", ".join(all_chart_types) if all_chart_types else "无"
    # 图形格式：包含图表类型和图片文件名
    if all_images:
        image_text = ", ".join(all_images)
        chart_display = f"{chart_type_text} ({image_text})" if chart_type_text != "无" else image_text
    else:
        chart_display = chart_type_text

    return {
        "sql": sql_text,
        "chart_type": chart_type_text,
        "chart_display": chart_display,
        "images": all_images,
    }


def save_result_xlsx(results: list, save_path: str):
    """保存结果为Excel文件（附件7表3格式）"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result_2"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D4D4D4'),
        right=Side(style='thin', color='D4D4D4'),
        top=Side(style='thin', color='D4D4D4'),
        bottom=Side(style='thin', color='D4D4D4'),
    )

    # 表头
    headers = ["编号", "问题", "SQL查询语句", "图形格式", "回答"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["id"]).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row_idx, column=2, value=json.dumps(r["questions"], ensure_ascii=False)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=3, value=r.get("sql", "")).alignment = wrap_alignment
        ws.cell(row=row_idx, column=4, value=r.get("chart_display", r.get("chart_type", "无"))).alignment = wrap_alignment
        ws.cell(row=row_idx, column=5, value=json.dumps(r["answers"], ensure_ascii=False)).alignment = wrap_alignment

        # 设置边框
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 90

    wb.save(save_path)
    logger.info(f"结果已保存: {save_path}")


async def main():
    start_time = time.time()

    print("=" * 70)
    print("  任务二：智能问数助手 - 回答附件4的问题")
    print("  开始时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 70)

    # 加载配置和初始化
    config = AppConfig.load()
    db = DatabaseManager(config.db_path)
    preflight = run_task2_preflight(config, db)
    emit_preflight_report(preflight, logger_instance=logger)
    if not preflight.ok:
        raise SystemExit("任务二运行前自检未通过，请先处理上述错误。")
    llm = LLMClient(config)

    # 检查数据库
    tables = db.get_table_names()
    if "income_sheet" not in tables:
        print("错误：数据库未初始化，请先运行 task1/run_task1.py")
        return

    row_count = db.get_table_row_count("income_sheet")
    print(f"\n数据库状态: income_sheet 有 {row_count} 条记录")

    # 初始化Agent
    results_dir = str(RESULTS_DIR)
    os.makedirs(results_dir, exist_ok=True)

    agent = SmartQAAgent(db=db, llm=llm, results_dir=results_dir)
    print("智能问答Agent初始化完成\n")

    # 加载问题
    questions = load_questions(SAMPLE_DATA_DIR)
    if not questions:
        print("未找到问题文件，使用示例问题")
        questions = [
            {"id": "B1001", "type": "数据基本查询", "questions": [{"Q": "金花股份利润总额是多少"}, {"Q": "2025年第三季度的"}]},
            {"id": "B1002", "type": "数据统计分析查询", "questions": [{"Q": "金花股份近几年的利润总额变化趋势是什么样的"}]},
        ]

    print(f"共 {len(questions)} 道问题待处理\n")

    # 逐个处理
    all_results = []
    for idx, q in enumerate(questions):
        print(f"\n{'─' * 60}")
        print(f"[{idx+1}/{len(questions)}] 问题 [{q['id']}] ({q['type']})")
        print(f"  内容: {json.dumps(q['questions'], ensure_ascii=False)}")

        # 处理多轮对话
        answers = await agent.process_conversation(
            questions=q["questions"],
            enhanced_mode=False,
            question_id=q["id"],
        )

        # 从answers中提取SQL、图表类型、图片路径
        extracted = _extract_results_from_answers(answers)

        result = {
            "id": q["id"],
            "questions": q["questions"],
            "answers": answers,
            "sql": extracted["sql"],
            "chart_type": extracted["chart_type"],
            "chart_display": extracted["chart_display"],
            "images": extracted["images"],
        }
        all_results.append(result)

        # 打印结果摘要
        print(f"  SQL: {extracted['sql'][:120]}..." if len(extracted['sql']) > 120 else f"  SQL: {extracted['sql'] or '无'}")
        print(f"  图表: {extracted['chart_display']}")
        for a in answers:
            content = a.get("A", {}).get("content", "")
            print(f"  回答: {content[:100]}...")

    # 保存结果
    result_path = str(RESULTS_DIR / "result_2.xlsx")
    save_result_xlsx(all_results, result_path)

    # 同时保存JSON格式（完整结果便于调试）
    json_path = str(RESULTS_DIR / "result_2.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f"\nJSON结果: {json_path}")

    # 打印统计
    total_sql = sum(1 for r in all_results if r["sql"])
    total_charts = sum(1 for r in all_results if r["chart_type"] != "无")
    total_images = sum(len(r.get("images", [])) for r in all_results)
    print(f"\n{'=' * 70}")
    print(f"  任务二完成！耗时: {time.time() - start_time:.1f} 秒")
    print(f"  结果文件: {result_path}")
    print(f"  SQL语句: {total_sql}/{len(all_results)} 道题")
    print(f"  图表生成: {total_charts} 个 ({total_images} 张图片)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    os.makedirs(str(PROJECT_ROOT / "logs"), exist_ok=True)
    os.makedirs(str(RESULTS_DIR), exist_ok=True)
    asyncio.run(main())
